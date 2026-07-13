"""
import_engine.py — CSV + Excel parser for the dynamic master-data import feature.

Task 3: parse_table(filename, content) -> (headers, rows)
  - CSV: utf-8-sig / cp1252 / latin-1 encoding fallback
  - Excel (.xlsx/.xlsm): openpyxl read_only + data_only mode
"""
import csv
import datetime as _dt
import io
import re as _re
from difflib import SequenceMatcher

from openpyxl import load_workbook

from field_registry import CONTROL_KEYS, control_key_for, list_fields, normalize_header


# ---------------------------------------------------------------------------
# P2.1 / P2.2 — text-safe cell coercion + phone normalization
# ---------------------------------------------------------------------------

# Scientific-notation detector. Excel displays large numbers as "9.17709E+11";
# when a CSV is saved that way the extra digits are already gone, so such cells
# are LOSSY and must be flagged, never silently stored.
_SCI_RE = _re.compile(r"[eE][+-]?\d")


def coerce_cell(v):
    """Return a text-safe string for a spreadsheet cell value.

    - integral floats/ints -> plain digits ("917709261234", never "9.17709E+11"
      or "917709261234.0")
    - date/datetime -> ISO ("2020-05-01" / "2020-05-01T00:00:00")
    - everything else -> str(v).strip()

    This keeps phone/id columns text-safe so long integers are never rendered in
    scientific notation or with a trailing ".0".
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        # avoid float repr / sci-notation for whole numbers
        if v.is_integer():
            return str(int(v))
        return repr(v)
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat()
    return str(v).strip()


def normalize_phone(raw) -> str:
    """Canonical phone string: keep digits, preserve a single leading '+',
    drop spaces / dashes / parens / a trailing '.0' from float coercion.

    Examples::
        "917709261234.0"    -> "917709261234"
        "+91 97709 12345"   -> "+919770912345"
        "(0120) 421-3000"   -> "01204213000"
        917709261234.0      -> "917709261234"
    """
    if raw is None:
        return ""
    # coerce numeric inputs first so 917709261234.0 doesn't keep the trailing 0
    if isinstance(raw, (int, float)):
        raw = coerce_cell(raw)
    s = str(raw).strip()
    if not s:
        return ""
    # drop a trailing ".0"/".00" left by float coercion before stripping non-digits
    s = _re.sub(r"\.0+$", "", s)
    plus = s.lstrip().startswith("+")
    digits = _re.sub(r"\D", "", s)
    if not digits:
        return ""
    return ("+" + digits) if plus else digits


def phone_is_lossy(raw) -> bool:
    """True if a raw phone value carries scientific notation (e.g. "9.17709E+11").

    Such values have already lost digits and cannot be recovered — the caller
    must flag them rather than normalizing garbage into `phone_norm`.
    """
    if raw is None or isinstance(raw, (int, float)):
        return False
    return bool(_SCI_RE.search(str(raw)))


def parse_table(filename: str, content: bytes) -> tuple[list[str], list[dict]]:
    """Parse CSV or Excel bytes into (headers, rows).

    Args:
        filename: Original filename — extension determines format.
        content:  Raw file bytes.

    Returns:
        headers: List of column name strings (stripped).
        rows:    List of dicts mapping header -> cell string (stripped).
                 Blank/all-None rows are skipped for Excel.
    """
    name = (filename or "").lower()

    if name.endswith((".xlsx", ".xlsm")):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = [
                str(h).strip() if h is not None else ""
                for h in next(rows_iter, [])
            ]
            rows = []
            for r in rows_iter:
                if r is None or all(c is None for c in r):
                    continue
                rows.append({
                    headers[i]: coerce_cell(v)
                    for i, v in enumerate(r)
                    if i < len(headers) and headers[i]
                })
            return headers, rows
        finally:
            wb.close()

    # CSV with encoding fallback (mirrors crm_routes.py encoding pattern)
    text = None
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue

    reader = csv.DictReader(io.StringIO(text or ""))
    headers = [h.strip() for h in (reader.fieldnames or [])]
    rows = [
        {(k or "").strip(): (v or "").strip() for k, v in row.items()}
        for row in reader
    ]
    return headers, rows


# ---------------------------------------------------------------------------
# Task 4: Auto column-mapping via alias lookup + fuzzy fallback
# ---------------------------------------------------------------------------
async def propose_mapping(db, headers: list) -> list:
    """Map raw spreadsheet headers to registered field definitions.

    For each header in *headers* (order preserved) returns a dict::

        {source, field_id, key, confidence}

    Confidence levels:
      "high"   — exact match on a normalized alias or label
      "medium" — best fuzzy SequenceMatcher ratio >= 0.78 across label+aliases
      "none"   — no match found
    """
    fields = await list_fields(db)

    # Build alias index: normalized alias/label -> field dict
    alias_index: dict = {}
    for f in fields:
        for a in f.get("aliases", []):
            alias_index[normalize_header(a)] = f
        # setdefault so an explicit alias wins over the derived label key
        alias_index.setdefault(normalize_header(f["label"]), f)

    out = []
    for h in headers:
        nh = normalize_header(h)

        # --- id control columns (round-trip / re-upload by ID) ---
        # school_id / contact_id / lead_id are not registered fields, but must
        # flow through so resolve_* can match by id and re-uploads update in
        # place instead of duplicating.
        ck = control_key_for(h)
        if ck:
            out.append({"source": h, "field_id": None, "key": ck, "confidence": "high"})
            continue

        # --- exact match ---
        f = alias_index.get(nh)
        if f:
            out.append({"source": h, "field_id": f["field_id"], "key": f["key"], "confidence": "high"})
            continue

        # --- fuzzy match ---
        best, score = None, 0.0
        for f2 in fields:
            candidates = [normalize_header(f2["label"])] + [normalize_header(a) for a in f2.get("aliases", [])]
            s = max(SequenceMatcher(None, nh, c).ratio() for c in candidates)
            if s > score:
                best, score = f2, s

        if best and score >= 0.78:
            out.append({"source": h, "field_id": best["field_id"], "key": best["key"], "confidence": "medium"})
        else:
            out.append({"source": h, "field_id": None, "key": None, "confidence": "none"})

    return out


# ---------------------------------------------------------------------------
# Task 5: Row resolver — school identity match
# ---------------------------------------------------------------------------
async def resolve_school(db, values: dict) -> dict:
    """Match a mapped row dict to an existing school record.

    Match precedence:
      1. values["school_id"] present and non-deleted school exists → update
      2. school_name (case-insensitive exact), optionally narrowed by city
         1 match → update, ≥2 → needs_review
      3. school_phone / phone exact match
         1 match → update, ≥2 → needs_review
      4. no match → create

    Returns:
        {"action": "create"|"update"|"needs_review", "school_id": str|None, "candidates": int}
    """
    sid = (values.get("school_id") or "").strip()
    if sid:
        hit = await db.schools.find_one(
            {"school_id": sid, "is_deleted": {"$ne": True}},
            {"_id": 0, "school_id": 1},
        )
        if hit:
            return {"action": "update", "school_id": sid, "candidates": 1}

    name = (values.get("school_name") or "").strip()
    city = (values.get("city") or "").strip()
    if name:
        q: dict = {
            "school_name": {"$regex": f"^{_re.escape(name)}$", "$options": "i"},
            "is_deleted": {"$ne": True},
            "school_id": {"$exists": True, "$ne": None},
        }
        if city:
            q["city"] = {"$regex": f"^{_re.escape(city)}$", "$options": "i"}
        cands = [d async for d in db.schools.find(q, {"_id": 0, "school_id": 1})]
        if len(cands) == 1:
            return {"action": "update", "school_id": cands[0].get("school_id"), "candidates": 1}
        if len(cands) >= 2:
            return {"action": "needs_review", "school_id": None, "candidates": len(cands)}

    phone = (values.get("school_phone") or values.get("phone") or "").strip()
    pn = normalize_phone(phone)
    if pn or phone:
        # Match on normalized phone (new rows) OR the raw phone (legacy rows).
        or_terms = []
        if pn:
            or_terms.append({"phone_norm": pn})
        if phone:
            or_terms.append({"phone": phone})
        cands = [d async for d in db.schools.find(
            {"$or": or_terms, "is_deleted": {"$ne": True},
             "school_id": {"$exists": True, "$ne": None}},
            {"_id": 0, "school_id": 1},
        )]
        if len(cands) == 1:
            return {"action": "update", "school_id": cands[0].get("school_id"), "candidates": 1}
        if len(cands) >= 2:
            return {"action": "needs_review", "school_id": None, "candidates": len(cands)}

    return {"action": "create", "school_id": None, "candidates": 0}


# Accept an externally-supplied id only if it is well-formed (safe id charset).
_ID_RE = _re.compile(r"^[A-Za-z0-9_-]{1,64}$")


def valid_supplied_id(raw) -> str:
    """Return a stripped, well-formed supplied id, or '' if blank/malformed.

    Honoring a supplied id lets a School/Contact/Lead ID round-trip on re-import
    instead of being discarded and re-minted."""
    s = str(raw or "").strip()
    return s if _ID_RE.match(s) else ""


# ---------------------------------------------------------------------------
# Task 6: split_values + commit_row — upsert School+Contact+Lead with audit
# ---------------------------------------------------------------------------
import uuid as _uuid

from field_registry import SEED_FIELDS


def _now() -> str:
    from datetime import datetime, timezone
    return datetime.now(timezone.utc).isoformat()


# Module-level map: key -> (entity, maps_to|None)
# Built from SEED_FIELDS tuple: (key, label, entity, type, maps_to, group, aliases)
_CORE = {key: (entity, maps_to) for (key, _l, entity, _t, maps_to, _g, _a) in SEED_FIELDS}


def split_values(row_keyed: dict) -> dict:
    """Split a mapped row dict into native and custom-field buckets per entity.

    Returns:
        {
          "school": {native_col: val, ...},
          "contact": {native_col: val, ...},
          "lead": {native_col: val, ...},
          "custom": {"school": {key: val}, "contact": {key: val}, "lead": {key: val}},
        }

    The control key "school_id" is skipped (it is used only for resolve_school).
    Keys whose maps_to is set write to native columns under their entity.
    Keys with no maps_to (or unknown keys) write to custom[entity][key].
    """
    out: dict = {
        "school": {},
        "contact": {},
        "lead": {},
        "custom": {"school": {}, "contact": {}, "lead": {}},
    }
    for key, val in row_keyed.items():
        if key in CONTROL_KEYS:         # control keys — not stored as fields
            continue
        meta = _CORE.get(key)
        if meta and meta[1]:            # known key with a native column mapping
            entity, native = meta
            out[entity][native] = val
        else:                           # custom field — goes into custom_fields
            entity = meta[0] if meta else "school"
            out["custom"][entity][key] = val
    return out


async def _resolve_import_owner(db, raw: str):
    """Resolve an import row's owner value (name OR email) to (email, name).

    Reuses crm_routes._resolve_owner logic (lazy import to avoid load-order
    coupling) so import ownership is keyed by EMAIL exactly like the CSV
    importer. A NAME is never stored in `assigned_to`."""
    raw = (raw or "").strip()
    if not raw:
        return "", ""
    from routes.crm_routes import resolve_owner
    return await resolve_owner(db, raw)


def _phone_pair(bucket: dict, entity: str, warnings: list) -> tuple[str, str]:
    """Return (raw_to_store, phone_norm) for a phone in an entity bucket.

    Lossy scientific-notation values are flagged in *warnings* and kept raw
    (for manual repair) with an EMPTY phone_norm so they never dedup falsely."""
    raw = bucket.get("phone", "") or ""
    if not raw:
        return "", ""
    if phone_is_lossy(raw):
        warnings.append({
            "entity": entity, "field": "phone", "value": raw,
            "message": "scientific-notation phone is lossy; stored raw, not normalized",
        })
        return raw, ""
    return raw, normalize_phone(raw)


async def commit_row(db, row_keyed: dict, user: dict, create_leads: bool) -> dict:
    """Upsert one import row into schools / contacts / leads with a pre-update audit snapshot.

    Safety rules:
    - If resolve_school returns needs_review → return immediately, write NOTHING.
    - owner: `assigned_to` is resolved name→email (P2.3); a name is never stored
      in `assigned_to`, only in `assigned_name`.
    - phones: stored normalized in `phone_norm` (P2.2); lossy sci-notation values
      are flagged and kept raw with a blank `phone_norm`.
    - dates: `import_date` on every write; `assigned_date` whenever a real owner
      email is set (P2.4).
    - ids: a well-formed supplied school_id / contact_id / lead_id is honored on
      create (round-trip), and matched first on re-import (P2.5).
    - create: mint sch_<12hex> only when no valid id supplied.
    - update: snapshot the prior school doc into audit_backup (kind=school_pre_import)
              BEFORE any $set.

    Returns:
        {"action", "school_id", "contact_id", "lead_id", "warnings"}
    """
    warnings: list = []
    user_email = user.get("email", "import")
    now = _now()

    parts = split_values(row_keyed)

    # --- P2.3 owner name→email (never store a name in assigned_to) ---
    raw_owner = str(parts["lead"].get("assigned_to") or "").strip()
    owner_email, owner_name = await _resolve_import_owner(db, raw_owner)
    if owner_email:
        assign_set = {"assigned_to": owner_email, "assigned_name": owner_name,
                      "assigned_date": now}
    elif owner_name:
        assign_set = {"assigned_name": owner_name}   # label only — scoping safe
    else:
        assign_set = {}

    # --- P2.1/P2.2 phone normalization + lossy flagging ---
    sch_phone_raw, sch_phone_norm = _phone_pair(parts["school"], "school", warnings)
    con_phone_raw, con_phone_norm = _phone_pair(parts["contact"], "contact", warnings)

    # --- P2.5 school resolution (id → name+city → phone_norm) ---
    res = await resolve_school(db, row_keyed)
    if res["action"] == "needs_review":
        return {"action": "needs_review", "school_id": None, "contact_id": None,
                "lead_id": None, "warnings": warnings}

    sid = res["school_id"]

    if res["action"] == "create":
        sid = valid_supplied_id(row_keyed.get("school_id")) or f"sch_{_uuid.uuid4().hex[:12]}"
        school_vals = dict(parts["school"])
        if sch_phone_raw or "phone" in school_vals:
            school_vals["phone"] = sch_phone_raw
        doc = {
            "school_id": sid,
            "is_deleted": False,
            "created_by": user_email,
            "created_at": now,
            "import_date": now,
            "custom_fields": parts["custom"]["school"],
            **school_vals,
        }
        if sch_phone_norm:
            doc["phone_norm"] = sch_phone_norm
        doc.update(assign_set)
        await db.schools.insert_one(doc)
    else:
        # Snapshot existing doc before overwriting (safety-critical — never skip)
        old = await db.schools.find_one({"school_id": sid})
        await db.audit_backup.insert_one({
            "kind": "school_pre_import",
            "school_id": sid,
            "snapshot": {k: v for k, v in (old or {}).items() if k != "_id"},
            "at": now,
            "by": user_email,
        })
        upd: dict = dict(parts["school"])
        if sch_phone_raw or "phone" in upd:
            upd["phone"] = sch_phone_raw
        for k, v in parts["custom"]["school"].items():
            upd[f"custom_fields.{k}"] = v
        upd["last_activity_date"] = now
        upd["import_date"] = now
        if sch_phone_norm:
            upd["phone_norm"] = sch_phone_norm
        upd.update(assign_set)
        await db.schools.update_one({"school_id": sid}, {"$set": upd})

    # ---- contact upsert: id → name+phone_norm → phone_norm → phone → name ----
    cid = None
    contact_name = parts["contact"].get("name", "")
    supplied_cid = valid_supplied_id(row_keyed.get("contact_id"))
    if contact_name or con_phone_raw or supplied_cid:
        existing = None
        if supplied_cid:
            existing = await db.contacts.find_one({"contact_id": supplied_cid})
        if existing is None and contact_name and con_phone_norm:
            existing = await db.contacts.find_one(
                {"school_id": sid, "name": contact_name, "phone_norm": con_phone_norm})
        if existing is None and con_phone_norm:
            existing = await db.contacts.find_one({"school_id": sid, "phone_norm": con_phone_norm})
        if existing is None and con_phone_raw:
            existing = await db.contacts.find_one({"school_id": sid, "phone": con_phone_raw})
        if existing is None and contact_name:
            existing = await db.contacts.find_one({"school_id": sid, "name": contact_name})

        cvals = dict(parts["contact"])
        cvals["phone"] = con_phone_raw
        cdoc = {
            "school_id": sid,
            **cvals,
            "phone_norm": con_phone_norm,
            "import_date": now,
            "custom_fields": parts["custom"]["contact"],
        }
        cdoc.update(assign_set)
        if existing:
            cid = existing["contact_id"]
            await db.contacts.update_one({"contact_id": cid}, {"$set": cdoc})
        else:
            cid = supplied_cid or f"con_{_uuid.uuid4().hex[:12]}"
            await db.contacts.insert_one({
                "contact_id": cid,
                "created_at": now,
                "is_deleted": False,
                "created_by": user_email,
                "status": "active",
                "company": parts["school"].get("school_name", ""),
                "source": "import",
                "converted_to_lead": False,
                "lead_id": None,
                **cdoc,
            })

    # ---- lead upsert: id → school; only if create_leads AND a real owner email ----
    lid = None
    supplied_lid = valid_supplied_id(row_keyed.get("lead_id"))
    if create_leads and owner_email:
        existing_lead = None
        if supplied_lid:
            existing_lead = await db.leads.find_one({"lead_id": supplied_lid})
        if existing_lead is None:
            existing_lead = await db.leads.find_one(
                {"school_id": sid, "is_deleted": {"$ne": True}})
        lead_set = {
            "assigned_to": owner_email,
            "assigned_name": owner_name,
            "assigned_date": now,
            "import_date": now,
        }
        if existing_lead:
            lid = existing_lead["lead_id"]
            await db.leads.update_one({"lead_id": lid}, {"$set": lead_set})
        else:
            lid = supplied_lid or f"lead_{_uuid.uuid4().hex[:12]}"
            await db.leads.insert_one({
                "lead_id": lid,
                "school_id": sid,
                "company_name": parts["school"].get("school_name", ""),
                "stage": "new",
                "created_at": now,
                "is_deleted": False,
                **lead_set,
            })

    return {"action": res["action"], "school_id": sid, "contact_id": cid,
            "lead_id": lid, "warnings": warnings}
