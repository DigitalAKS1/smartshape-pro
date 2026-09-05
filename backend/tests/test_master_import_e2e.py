"""End-to-end Import Center round-trip, through the real HTTP handlers.

The owner's workflow, exactly: hit Export all (.xlsx), open the workbook, edit /
rearrange / delete columns, upload it back through preview, commit through
execute. test_master_export_roundtrip.py covers the helpers; this covers the
route path and what the owner is actually TOLD afterwards. mongomock.
"""
import asyncio
import io
import os

os.environ.setdefault("MONGO_URL", "mongodb://localhost:27017")
os.environ.setdefault("DB_NAME", "smartshape_test")

import pytest
from mongomock_motor import AsyncMongoMockClient
from openpyxl import Workbook, load_workbook

import field_registry as fr
import routes.dynamic_import_routes as dyn

ADMIN = {"email": "info@smartshape.in", "name": "Owner", "role": "admin"}


class FakeUpload:
    """Stands in for a Starlette UploadFile."""

    def __init__(self, data, filename="school-master-export.xlsx"):
        self._data = data
        self.filename = filename

    async def read(self):
        return self._data


@pytest.fixture()
def db(monkeypatch):
    d = AsyncMongoMockClient()["smartshape_test"]
    monkeypatch.setattr(dyn, "db", d)
    import import_engine
    monkeypatch.setattr(import_engine, "db", d, raising=False)
    return d


def _run(coro):
    return asyncio.run(coro)


async def _seed(db):
    """Two schools. One has three contacts, one has none — both real shapes."""
    await fr.seed_field_definitions(db)
    await db.users.insert_one(
        {"email": "parul@smartshape.in", "name": "Parul Kanchan", "is_active": True, "role": "sales"})
    await db.schools.insert_one({
        "school_id": "sch_dps", "school_name": "Delhi Public School",
        "school_type": "CBSE", "city": "Rohini", "state": "Delhi",
        "phone": "01127654321", "email": "info@dps.in", "school_strength": 1200,
        "board": "CBSE", "website": "https://dps.in",
        "assigned_to": "parul@smartshape.in", "assigned_name": "Parul Kanchan",
        "custom_fields": {"std_classes": "1-12"},
        "is_deleted": False,
    })
    for cid, name, desig, phone in (
        ("con_p", "R Sharma", "Principal", "9811111111"),
        ("con_d", "K Verma", "Director", "9822222222"),
        ("con_c", "S Gupta", "Coordinator", "9833333333"),
    ):
        await db.contacts.insert_one({
            "contact_id": cid, "school_id": "sch_dps", "name": name,
            "designation": desig, "phone": phone, "is_deleted": False,
        })
    await db.schools.insert_one({
        "school_id": "sch_lone", "school_name": "Lotus Valley",
        "city": "Noida", "school_strength": 400, "is_deleted": False,
    })


async def _export_workbook(db):
    """Call the real .xlsx endpoint and hand back the parsed workbook bytes."""
    resp = await dyn.master_export_xlsx(user=ADMIN)
    return resp.body


def _edit_workbook(content, edit):
    """Open exported bytes, run `edit(ws, headers)`, hand back new bytes."""
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    edit(ws, headers)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _reupload(db, content):
    preview = await dyn.import_preview(file=FakeUpload(content), entity_type="school", user=ADMIN)
    result = await dyn.import_execute({
        "rows_keyed": preview["rows_keyed"],
        "mapping": preview["mapping"],
        "create_leads": False,
    }, user=ADMIN)
    return preview, result


# ── The whole loop ──────────────────────────────────────────────────────────

def test_export_edit_reupload_applies_the_edit_and_creates_nothing(db):
    async def go():
        await _seed(db)
        content = await _export_workbook(db)

        def edit(ws, headers):
            city = headers.index("City") + 1
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, headers.index("School ID") + 1).value == "sch_dps":
                    ws.cell(r, city).value = "Pitampura"

        _, result = await _reupload(db, _edit_workbook(content, edit))

        assert result["counts"]["error"] == 0, result["errors"]
        assert result["counts"]["create"] == 0, "a re-upload must never create a school"
        assert await db.schools.count_documents({}) == 2
        assert await db.contacts.count_documents({}) == 3
        school = await db.schools.find_one({"school_id": "sch_dps"})
        assert school["city"] == "Pitampura"
    _run(go())


def test_the_owner_is_told_how_many_SCHOOLS_changed_not_how_many_rows(db):
    # The export writes one row per contact, so a 2-school file is 4 rows. If the
    # summary counts rows, "4 updated" for 2 schools reads as data being
    # duplicated and there is no way to tell that it wasn't.
    async def go():
        await _seed(db)
        _, result = await _reupload(db, await _export_workbook(db))
        assert result["total"] == 4, "fixture should produce 3 contact rows + 1 lone school"
        assert result["counts"]["update"] == 2, \
            f"reported {result['counts']['update']} updates for 2 schools"
    _run(go())


def test_every_contact_survives_the_loop_and_edits_reach_the_right_one(db):
    async def go():
        await _seed(db)
        content = await _export_workbook(db)

        def edit(ws, headers):
            cid = headers.index("Contact ID") + 1
            name = headers.index("Name") + 1
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, cid).value == "con_d":
                    ws.cell(r, name).value = "Kavita Verma"

        await _reupload(db, _edit_workbook(content, edit))

        assert await db.contacts.count_documents({}) == 3, "a rename must not fork the contact"
        names = {c["contact_id"]: c["name"] async for c in db.contacts.find({})}
        assert names == {"con_p": "R Sharma", "con_d": "Kavita Verma", "con_c": "S Gupta"}
    _run(go())


def test_reuploading_the_same_file_twice_changes_nothing_further(db):
    async def go():
        await _seed(db)
        content = await _export_workbook(db)
        await _reupload(db, content)
        first = await db.schools.find_one({"school_id": "sch_dps"}, {"_id": 0})
        await _reupload(db, content)
        second = await db.schools.find_one({"school_id": "sch_dps"}, {"_id": 0})
        for volatile in ("import_date", "last_activity_date", "assigned_date"):
            first.pop(volatile, None)
            second.pop(volatile, None)
        assert first == second
        assert await db.schools.count_documents({}) == 2
        assert await db.contacts.count_documents({}) == 3
    _run(go())


# ── Rearranging, which is the half the owner asked about ────────────────────

def test_reordered_columns_still_map(db):
    async def go():
        await _seed(db)
        wb = load_workbook(io.BytesIO(await _export_workbook(db)))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
        order = list(reversed(range(len(headers))))   # every column moved

        out = Workbook()
        os_ = out.active
        os_.append([headers[i] for i in order])
        for row in rows:
            os_.append([row[i] for i in order])
        buf = io.BytesIO()
        out.save(buf)

        _, result = await _reupload(db, buf.getvalue())
        assert result["counts"]["error"] == 0, result["errors"]
        assert result["counts"]["create"] == 0
        assert (await db.schools.find_one({"school_id": "sch_dps"}))["city"] == "Rohini"
    _run(go())


def test_a_deleted_column_leaves_that_field_alone_instead_of_wiping_it(db):
    # Trimming the sheet to the few columns you want to edit is the normal way
    # to use a spreadsheet. It must not read as "blank out everything else".
    async def go():
        await _seed(db)
        wb = load_workbook(io.BytesIO(await _export_workbook(db)))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        keep = ["School ID", "City"]
        idx = [headers.index(h) for h in keep]
        rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]

        out = Workbook()
        os_ = out.active
        os_.append(keep)
        for row in rows:
            os_.append([row[i] for i in idx])
        buf = io.BytesIO()
        out.save(buf)

        _, result = await _reupload(db, buf.getvalue())
        assert result["counts"]["error"] == 0, result["errors"]
        school = await db.schools.find_one({"school_id": "sch_dps"})
        assert school["board"] == "CBSE", "an absent column wiped a field"
        assert school["school_name"] == "Delhi Public School"
        assert school["school_strength"] == 1200
    _run(go())


def test_an_unrecognised_extra_column_is_ignored_not_fatal(db):
    async def go():
        await _seed(db)
        wb = load_workbook(io.BytesIO(await _export_workbook(db)))
        ws = wb.active
        ws.cell(1, ws.max_column + 1).value = "Notes from the sales meeting"
        for r in range(2, ws.max_row + 1):
            ws.cell(r, ws.max_column).value = "call back in May"
        buf = io.BytesIO()
        wb.save(buf)

        preview, result = await _reupload(db, buf.getvalue())
        assert result["counts"]["error"] == 0, result["errors"]
        unmapped = [m["source"] for m in preview["mapping"] if not m.get("key")]
        assert "Notes from the sales meeting" in unmapped, \
            "an unknown column should be reported as unmapped, not guessed at"
    _run(go())


def test_clearing_a_cell_clears_the_field(db):
    # The flip side of the rule above: a column that IS present and blank is an
    # instruction to empty that field. Documented here so it can't drift.
    async def go():
        await _seed(db)
        content = await _export_workbook(db)

        def edit(ws, headers):
            col = headers.index("School Website") + 1
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).value = None

        await _reupload(db, _edit_workbook(content, edit))
        assert (await db.schools.find_one({"school_id": "sch_dps"}))["website"] == ""
    _run(go())


# ── The owner column, and the reassign review panel ─────────────────────────

def test_editing_the_owner_name_lists_the_school_once_for_review(db):
    async def go():
        await _seed(db)
        await db.users.insert_one(
            {"email": "amit@smartshape.in", "name": "Amit Rao", "is_active": True, "role": "sales"})
        content = await _export_workbook(db)

        def edit(ws, headers):
            col = headers.index("Assign To") + 1
            sid = headers.index("School ID") + 1
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, sid).value == "sch_dps":
                    ws.cell(r, col).value = "Amit Rao"

        preview = await dyn.import_preview(
            file=FakeUpload(_edit_workbook(content, edit)), entity_type="school", user=ADMIN)
        plans = [r for r in preview["reassignments"] if r.get("school_id") == "sch_dps"]
        assert len(plans) == 1, \
            f"the review panel lists sch_dps {len(plans)} times — once per contact row"
        assert plans[0]["to_email"] == "amit@smartshape.in"
    _run(go())


def test_an_owner_name_that_matches_nobody_is_flagged_not_silently_dropped(db):
    async def go():
        await _seed(db)
        content = await _export_workbook(db)

        def edit(ws, headers):
            col = headers.index("Assign To") + 1
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).value = "Someone Who Left"

        preview = await dyn.import_preview(
            file=FakeUpload(_edit_workbook(content, edit)), entity_type="school", user=ADMIN)
        flagged = [r for r in preview["reassignments"] if r["status"] == "owner_unmatched"]
        assert flagged, "an unresolvable owner must reach the review panel"
        school = await db.schools.find_one({"school_id": "sch_dps"})
        assert school["assigned_to"] == "parul@smartshape.in", "preview must not write"
    _run(go())


# ── Adding rows, which is the other half of "update and reupload" ───────────

def test_a_brand_new_row_with_no_School_ID_creates_a_school(db):
    async def go():
        await _seed(db)
        wb = load_workbook(io.BytesIO(await _export_workbook(db)))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        new = [""] * len(headers)
        new[headers.index("School/Institute Name")] = "Ryan International"
        new[headers.index("City")] = "Gurgaon"
        new[headers.index("Name")] = "A Menon"
        new[headers.index("Phone Number")] = "9844444444"
        ws.append(new)
        buf = io.BytesIO()
        wb.save(buf)

        _, result = await _reupload(db, buf.getvalue())
        assert result["counts"]["error"] == 0, result["errors"]
        assert result["counts"]["create"] == 1
        created = await db.schools.find_one({"school_name": "Ryan International"})
        assert created["city"] == "Gurgaon"
        assert await db.contacts.count_documents({"school_id": created["school_id"]}) == 1
    _run(go())
