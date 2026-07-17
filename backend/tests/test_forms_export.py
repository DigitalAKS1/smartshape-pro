"""test_forms_export.py — responses listing + CSV/XLSX export."""
import csv, io, os
import pytest
import pytest_asyncio
from motor.motor_asyncio import AsyncIOMotorClient
from fastapi import FastAPI
from httpx import AsyncClient, ASGITransport

_DB_NAME = os.getenv("DB_NAME", "smartshape_test")
assert _DB_NAME.endswith("_test") or _DB_NAME == "mtt_ci", f"refusing non-test DB: {_DB_NAME}"

import routes.form_routes as fr
import routes.training_routes as tr

_MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017")
SALES = {"email": "rep@smartshape.in", "role": "sales_person", "module_permissions": {}}
OTHER = {"email": "other@smartshape.in", "role": "sales_person", "module_permissions": {}}


@pytest_asyncio.fixture
async def ctx(monkeypatch):
    motor_client = AsyncIOMotorClient(_MONGO_URL)
    d = motor_client[_DB_NAME]
    orig_db = fr.db
    orig_tr = tr.db
    fr.db = d
    tr.db = d          # _enqueue_webinar_stage writes via training_routes.db
    fr._RATE.clear()
    fr._CURRENT = {"user": SALES}
    async def fake_user(request):
        return fr._CURRENT["user"]
    monkeypatch.setattr(fr, "get_current_user", fake_user)
    app = FastAPI()
    app.include_router(fr.router, prefix="/api")
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://t") as client:
        form = (await client.post("/api/forms", json={
            "title": "Session #05", "type": "event",
            "event": {"theme": "Patriotism", "date": "2026-07-18", "time": "13:00",
                      "meeting_link": "https://zoom.us/j/999"}})).json()
        by_map = {f["map_to"]: f["field_id"] for f in form["fields"] if f.get("map_to")}
        ans = {by_map[k]: v for k, v in {
            "name": "Asha Verma", "email": "asha@example.com", "school": "DPS Indore",
            "designation": "PRT", "phone": "9876543210", "city": "Indore"}.items()}
        await client.post(f"/api/forms/public/{form['public_token']}/submit",
                          json={"answers": ans})
        yield d, client, form
    fr.db, tr.db = orig_db, orig_tr
    for coll in ("forms", "form_responses", "training_sessions", "session_registrations",
                 "email_scheduled", "whatsapp_scheduled", "email_campaigns",
                 "contacts", "schools"):
        await d[coll].delete_many({})
    motor_client.close()


@pytest.mark.asyncio
async def test_responses_listing_scoped(ctx):
    d, client, form = ctx
    r = await client.get(f"/api/forms/{form['form_id']}/responses")
    assert r.status_code == 200
    body = r.json()
    assert body["count"] == 1
    assert body["responses"][0]["delivery"]["email"] == "queued"
    fr._CURRENT["user"] = OTHER
    assert (await client.get(f"/api/forms/{form['form_id']}/responses")).status_code == 403
    fr._CURRENT["user"] = SALES


@pytest.mark.asyncio
async def test_csv_export(ctx):
    d, client, form = ctx
    r = await client.get(f"/api/forms/{form['form_id']}/export.csv")
    assert r.status_code == 200
    assert "text/csv" in r.headers["content-type"]
    rows = list(csv.reader(io.StringIO(r.text)))
    assert rows[0][:2] == ["Submitted At", "Name"]
    assert "Asha Verma" in rows[1]


@pytest.mark.asyncio
async def test_xlsx_export(ctx):
    d, client, form = ctx
    r = await client.get(f"/api/forms/{form['form_id']}/export.xlsx")
    assert r.status_code == 200
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    ws = wb.active
    data = [[c.value for c in row] for row in ws.iter_rows()]
    assert data[0][1] == "Name" and "Asha Verma" in data[1]


@pytest.mark.asyncio
async def test_share_responses_emails_xlsx(ctx, monkeypatch):
    d, client, form = ctx
    import scheduler
    calls = {}
    async def fake_cfg():
        return ("info@smartshape.in", "app-pw", "SmartShape")
    def fake_send(cfg, recipients, subject, body, xlsx, filename):
        calls["recipients"] = recipients
        calls["filename"] = filename
        calls["xlsx_len"] = len(xlsx)
    monkeypatch.setattr(scheduler, "_email_cfg", fake_cfg)
    monkeypatch.setattr(fr, "_send_xlsx_email_sync", fake_send)
    # no recipients -> 422
    r0 = await client.post(f"/api/forms/{form['form_id']}/share-responses", json={"recipients": ""})
    assert r0.status_code == 422
    # valid recipients -> sends, dedups + parses list
    r = await client.post(f"/api/forms/{form['form_id']}/share-responses",
                          json={"recipients": "owner@smartshape.in, owner@smartshape.in team@x.co",
                                "note": "here you go"})
    assert r.status_code == 200
    assert r.json()["count"] == 1
    assert calls["recipients"] == ["owner@smartshape.in", "team@x.co"]
    assert calls["xlsx_len"] > 0 and calls["filename"].endswith(".xlsx")
    fr._CURRENT["user"] = OTHER
    assert (await client.post(f"/api/forms/{form['form_id']}/share-responses",
                             json={"recipients": "x@y.co"})).status_code == 403
    fr._CURRENT["user"] = SALES
